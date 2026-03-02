"""
Script para extrair os horários de trabalho de cada profissional por dia da semana.

Para cada profissional e cada dia da semana, identifica:
- O primeiro horário com qualquer célula preenchida (VAGO ou paciente) = entrada
- O último horário com qualquer célula preenchida = saída

Gera professional_schedules.csv com:
- professional_full_name
- day_of_week (0=segunda ... 4=sexta)
- day_of_week_name
- work_start (HH:MM) — primeiro horário do dia
- work_end (HH:MM) — fim do último horário do dia
- total_slots — total de slots no dia (vagos + com paciente)
- vacant_slots — slots sem paciente (VAGO)
- occupied_slots — slots com paciente
"""
import openpyxl
import csv
import re

DAYS = {
    'SEGUNDA-FEIRA': 0,
    'TERÇA-FEIRA': 1,
    'QUARTA-FEIRA': 2,
    'QUINTA-FEIRA': 3,
    'SEXTA-FEIRA': 4,
}

DAY_NAMES_PT = {
    0: 'segunda-feira',
    1: 'terça-feira',
    2: 'quarta-feira',
    3: 'quinta-feira',
    4: 'sexta-feira',
}

def parse_horario(horario_str):
    """Converte '08h00 - 08h50' em (h_start, m_start, h_end, m_end)"""
    if not horario_str:
        return None
    match = re.match(r'(\d{2})h(\d{2})\s*-\s*(\d{2})h(\d{2})', str(horario_str))
    if match:
        return int(match.group(1)), int(match.group(2)), int(match.group(3)), int(match.group(4))
    return None

def is_vago(cell_value):
    """Retorna True se a célula está vaga (sem paciente)."""
    if not cell_value:
        return False
    v = str(cell_value).strip().upper()
    return v.startswith('VAGO')

def is_occupied(cell_value):
    """Retorna True se a célula tem paciente ou atendimento."""
    if not cell_value:
        return False
    v = str(cell_value).strip().upper()
    skip = ['VAGO', 'SUPERVISÃO', 'SUPERVISAO', 'SAÍDA', 'SAIDA', 'FOLGA', 'FERIADO']
    if any(k in v for k in skip):
        return False
    return len(v) >= 3

def extract_professional_name(sheet_name):
    match = re.match(r'^(.+?)\s*[\(\(]', sheet_name)
    if match:
        return match.group(1).strip()
    return sheet_name.strip()

def process_sheet(ws, professional_name):
    """
    Retorna lista de dicts com horários de trabalho por dia da semana.
    """
    # Encontrar cabeçalho
    header_row = None
    day_columns = {}

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == 'SEGUNDA-FEIRA':
                header_row = cell.row
                for c in ws.iter_cols(min_row=header_row, max_row=header_row,
                                      min_col=1, max_col=ws.max_column):
                    for hcell in c:
                        day_name = str(hcell.value).strip() if hcell.value else ''
                        if day_name in DAYS:
                            day_columns[hcell.column] = DAYS[day_name]
                break
        if header_row:
            break

    if not header_row or not day_columns:
        return []

    horario_col = 1

    # Para cada dia: lista de (h_start, m_start, h_end, m_end, status)
    # status: 'vago' | 'occupied' | None (célula vazia)
    day_slots = {d: [] for d in day_columns.values()}

    for row in ws.iter_rows(min_row=header_row + 1):
        horario_cell = None
        for cell in row:
            if cell.column == horario_col:
                horario_cell = cell
                break

        if not horario_cell or not horario_cell.value:
            continue

        horario_parsed = parse_horario(str(horario_cell.value).strip())
        if not horario_parsed:
            continue

        h_start, m_start, h_end, m_end = horario_parsed

        for col_idx, day_of_week in day_columns.items():
            cell_value = None
            for cell in row:
                if cell.column == col_idx:
                    cell_value = cell.value
                    break

            if cell_value is None or str(cell_value).strip() == '':
                continue  # célula vazia = profissional não trabalha nesse slot

            if is_vago(cell_value):
                status = 'vago'
            elif is_occupied(cell_value):
                status = 'occupied'
            else:
                continue  # supervisão, saída, etc. — ignorar

            day_slots[day_of_week].append((h_start, m_start, h_end, m_end, status))

    # Montar resultado
    results = []
    for day_of_week, slots in day_slots.items():
        if not slots:
            continue  # profissional não trabalha nesse dia

        # Ordenar por horário de início
        slots_sorted = sorted(slots, key=lambda x: (x[0], x[1]))

        work_start_h, work_start_m = slots_sorted[0][0], slots_sorted[0][1]
        work_end_h, work_end_m = slots_sorted[-1][2], slots_sorted[-1][3]

        total_slots = len(slots_sorted)
        vacant_slots = sum(1 for s in slots_sorted if s[4] == 'vago')
        occupied_slots = sum(1 for s in slots_sorted if s[4] == 'occupied')

        results.append({
            'professional_full_name': professional_name,
            'day_of_week': day_of_week,
            'day_of_week_name': DAY_NAMES_PT[day_of_week],
            'work_start': f'{work_start_h:02d}:{work_start_m:02d}',
            'work_end': f'{work_end_h:02d}:{work_end_m:02d}',
            'total_slots': total_slots,
            'vacant_slots': vacant_slots,
            'occupied_slots': occupied_slots,
        })

    return results


def main():
    wb = openpyxl.load_workbook('/home/ubuntu/upload/AgendaProfissionais.xlsx', data_only=True)
    skip_sheets = {'Atendimentos', 'Cópia de Modelo'}

    all_schedules = []

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        professional_name = extract_professional_name(sheet_name)
        print(f"Processando: {professional_name}")

        schedules = process_sheet(ws, professional_name)
        print(f"  -> {len(schedules)} dias de trabalho encontrados")
        all_schedules.extend(schedules)

    print(f"\nTotal de registros: {len(all_schedules)}")

    # Salvar CSV
    output_path = '/home/ubuntu/agenda-aba/migration/professional_schedules.csv'
    fieldnames = ['professional_full_name', 'day_of_week', 'day_of_week_name',
                  'work_start', 'work_end', 'total_slots', 'vacant_slots', 'occupied_slots']

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_schedules)

    print(f"CSV salvo em: {output_path}")

    # Amostra
    print("\n--- Amostra dos primeiros 10 registros ---")
    for s in all_schedules[:10]:
        print(s)

    # Estatísticas
    total_vacant = sum(s['vacant_slots'] for s in all_schedules)
    total_occupied = sum(s['occupied_slots'] for s in all_schedules)
    total_slots = sum(s['total_slots'] for s in all_schedules)

    print(f"\n--- Estatísticas gerais ---")
    print(f"Total de slots na semana: {total_slots}")
    print(f"Slots com paciente: {total_occupied}")
    print(f"Slots vagos: {total_vacant}")
    print(f"Taxa de ocupação: {total_occupied/total_slots*100:.1f}%")

    # Por profissional
    print(f"\n--- Vagos por profissional ---")
    prof_vacant = {}
    for s in all_schedules:
        p = s['professional_full_name']
        prof_vacant[p] = prof_vacant.get(p, 0) + s['vacant_slots']
    for p, v in sorted(prof_vacant.items(), key=lambda x: -x[1]):
        print(f"  {p}: {v} slots vagos")


if __name__ == '__main__':
    main()
