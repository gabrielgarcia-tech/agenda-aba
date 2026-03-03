"""
Script para processar a planilha AgendaProfissionais.xlsx e gerar o CSV de migração
no formato de RECORRÊNCIA SEMANAL.

Cada linha representa uma regra recorrente:
- day_of_week: 0=Segunda, 1=Terça, 2=Quarta, 3=Quinta, 4=Sexta
- start_time: horário de início (HH:MM)
- end_time: horário de fim (HH:MM)
- professional_full_name: nome completo do profissional
- patient_full_name: nome completo do paciente
- appointment_type: clinic | school
- is_group_appointment: true | false
- notes: observações

Casos especiais tratados:
- Marcela Borilli e Larissa Thais: coluna extra de horários entre quinta e sexta (col 5=horários quinta, col 6=pacientes quinta, col 7=sexta)
- Atendimentos em dupla: "Dupla: Nome1 / Nome2" → duas linhas separadas com is_group_appointment=true
- Horários divergentes: "Das HHhMM às HHhMM Nome"
- Atendimentos escolares: "Atendimento Escolar: Nome"
- Sufixos de especialidade removidos: "- PSICO", "- FONO", etc.
"""
import openpyxl
from openpyxl.cell.cell import MergedCell
import csv
import re

DAYS = {
    'SEGUNDA-FEIRA': 0,
    'TERÇA-FEIRA': 1,
    'QUARTA-FEIRA': 2,
    'QUINTA-FEIRA': 3,
    'SEXTA-FEIRA': 4,
}

DAY_NAMES_PT = {
    0: 'segunda-feira',
    1: 'terça-feira',
    2: 'quarta-feira',
    3: 'quinta-feira',
    4: 'sexta-feira',
}

def parse_horario(horario_str):
    """Converte '08h00 - 08h50' em (h_start, m_start, h_end, m_end)"""
    if not horario_str:
        return None
    match = re.match(r'(\d{1,2})h(\d{2})\s*[-–]\s*(\d{1,2})h(\d{2})', str(horario_str))
    if match:
        return int(match.group(1)), int(match.group(2)), int(match.group(3)), int(match.group(4))
    return None

def clean_name(name):
    """Remove espaços extras, sufixos de especialidade e normaliza o nome."""
    if not name:
        return None
    name = str(name).strip()
    name = re.sub(r'\s*-\s*(PSICO|FONO|TO|ABA|PSICOLOGIA|FONOAUDIOLOGIA)\s*$', '', name, flags=re.IGNORECASE)
    return name.strip()

def extract_professional_name(sheet_name):
    match = re.match(r'^(.+?)\s*[\(\(]', sheet_name)
    if match:
        return match.group(1).strip()
    return sheet_name.strip()

def make_record(day_of_week, h_start, m_start, h_end, m_end, professional_name,
                patient_name, appointment_type='clinic', is_group=False, notes=''):
    return {
        'day_of_week': day_of_week,
        'day_of_week_name': DAY_NAMES_PT[day_of_week],
        'start_time': f'{h_start:02d}:{m_start:02d}',
        'end_time': f'{h_end:02d}:{m_end:02d}',
        'professional_full_name': professional_name,
        'patient_full_name': patient_name,
        'appointment_type': appointment_type,
        'is_group_appointment': 'true' if is_group else 'false',
        'notes': notes,
    }

def process_cell(cell_value, horario_parsed, day_of_week, professional_name):
    """
    Processa o conteúdo de uma célula e retorna lista de registros.
    """
    if not cell_value:
        return []

    value = str(cell_value).strip()

    # Ignorar células sem conteúdo útil
    skip_keywords = ['VAGO', 'SUPERVISÃO', 'SUPERVISAO', 'SAÍDA', 'SAIDA', 'FOLGA', 'FERIADO']
    if any(kw.lower() in value.lower() for kw in skip_keywords):
        return []

    # Ignorar se o valor parece um horário (ex: "08h00 - 08h50") — coluna extra de horários
    if re.match(r'^\d{1,2}h\d{2}\s*[-–]', value):
        return []

    if not horario_parsed:
        return []

    h_start, m_start, h_end, m_end = horario_parsed

    # Horário divergente: "Das HHhMM às HHhMM Nome" ou "HHhMM às HHhMM Nome"
    divergent_match = re.match(
        r'(?:Das?\s+)?(\d{1,2})h(\d{2})\s+[\àa]s?\s+(\d{1,2})h(\d{2})\s*[-–]?\s*(.+)',
        value, re.IGNORECASE
    )
    if divergent_match:
        h_start = int(divergent_match.group(1))
        m_start = int(divergent_match.group(2))
        h_end   = int(divergent_match.group(3))
        m_end   = int(divergent_match.group(4))
        patient_name = clean_name(divergent_match.group(5))
        if not patient_name:
            return []
        return [make_record(day_of_week, h_start, m_start, h_end, m_end,
                            professional_name, patient_name, notes='Horário divergente')]

    # Atendimento escolar: "Atendimento Escolar: Nome"
    school_match = re.match(r'Atendimento Escolar[:\s]+(.+)', value, re.IGNORECASE)
    if school_match:
        patient_name = clean_name(school_match.group(1))
        if not patient_name:
            return []
        return [make_record(day_of_week, h_start, m_start, h_end, m_end,
                            professional_name, patient_name,
                            appointment_type='school', notes='Atendimento Escolar')]

    # Atendimento em dupla: "Dupla: Nome1 / Nome2"
    dupla_match = re.match(r'Dupla[:\s]+(.+)', value, re.IGNORECASE)
    if dupla_match:
        nomes_raw = dupla_match.group(1)
        nomes = [clean_name(n) for n in re.split(r'\s*/\s*', nomes_raw)]
        nomes = [n for n in nomes if n and len(n) >= 3]
        records = []
        for nome in nomes:
            records.append(make_record(day_of_week, h_start, m_start, h_end, m_end,
                                       professional_name, nome,
                                       is_group=True, notes='Atendimento em dupla'))
        return records

    # Atendimento normal
    patient_name = clean_name(value)
    if not patient_name or len(patient_name) < 3:
        return []

    return [make_record(day_of_week, h_start, m_start, h_end, m_end,
                        professional_name, patient_name)]


def get_day_columns(ws):
    """
    Retorna dict {col_index: day_of_week} mapeando colunas de pacientes para dias.
    Lida com o caso especial de Marcela/Larissa onde há coluna extra de horários.
    """
    day_columns = {}  # col_index_de_pacientes -> day_of_week

    for row in ws.iter_rows():
        found_day = False
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value and str(cell.value).strip() in DAYS:
                found_day = True
                day_name = str(cell.value).strip()
                day_of_week = DAYS[day_name]
                col = cell.column

                # Verificar se a próxima coluna é também um dia da semana
                # Se não for, é a coluna de pacientes deste dia
                # Se a coluna seguinte for um horário (não um dia), então a coluna de pacientes
                # é col+1 (caso especial com coluna extra de horários)
                next_col_header = None
                for c2 in row:
                    if isinstance(c2, MergedCell):
                        continue
                    if c2.column == col + 1:
                        next_col_header = c2.value
                        break

                if next_col_header and str(next_col_header).strip() in DAYS:
                    # Próxima coluna é outro dia — pacientes estão na mesma coluna do cabeçalho
                    day_columns[col] = day_of_week
                elif next_col_header is None or str(next_col_header).strip() == '':
                    # Próxima coluna está vazia — pode ser coluna extra de horários
                    # Verificar se col+1 tem horários nas linhas de dados
                    # Heurística: se o cabeçalho do dia está em col e col+1 está vazio,
                    # os pacientes estão em col+1
                    # Mas se col+1 também não é dia da semana, usamos col+1 como pacientes
                    day_columns[col + 1] = day_of_week
                else:
                    day_columns[col] = day_of_week

        if found_day:
            break

    return day_columns


def process_sheet(ws, professional_name):
    """Processa uma aba e retorna lista de registros de recorrência."""
    appointments = []

    # Encontrar linha do cabeçalho
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value and str(cell.value).strip() == 'SEGUNDA-FEIRA':
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print(f"  [AVISO] Cabeçalho não encontrado: {professional_name}")
        return []

    # Mapear colunas de dias — usar a linha do cabeçalho encontrada dinamicamente
    day_columns = {}
    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    for cell in header_cells:
        if isinstance(cell, MergedCell):
            continue
        if cell.value and str(cell.value).strip() in DAYS:
            day_name = str(cell.value).strip()
            day_of_week = DAYS[day_name]
            col = cell.column

            # Verificar próxima coluna no cabeçalho
            next_val = None
            for c2 in header_cells:
                if isinstance(c2, MergedCell):
                    continue
                if c2.column == col + 1:
                    next_val = c2.value
                    break

            DAYS_SET = set(DAYS.keys())
            if next_val and str(next_val).strip() in DAYS_SET:
                # Próxima coluna é outro dia → pacientes na mesma coluna do cabeçalho
                day_columns[col] = day_of_week
            elif next_val is not None and str(next_val).strip() == '':
                # Próxima coluna está vazia → caso especial (coluna extra de horários)
                # Pacientes estão na col+1
                day_columns[col + 1] = day_of_week
            else:
                # Próxima coluna não existe (ultima coluna) ou tem outro conteúdo
                # Pacientes estão na mesma coluna do cabeçalho
                day_columns[col] = day_of_week

    # Coluna de horários (sempre col 1)
    horario_col = 1

    # Processar linhas de dados
    for row in ws.iter_rows(min_row=header_row + 1):
        # Pegar horário da coluna A
        horario_val = None
        row_vals = {}
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            row_vals[cell.column] = cell.value
            if cell.column == horario_col:
                horario_val = cell.value

        if not horario_val:
            continue

        horario_parsed = parse_horario(str(horario_val).strip())
        if not horario_parsed:
            continue

        # Processar cada coluna de dia
        for col_idx, day_of_week in day_columns.items():
            cell_value = row_vals.get(col_idx)
            records = process_cell(cell_value, horario_parsed, day_of_week, professional_name)
            appointments.extend(records)

    return appointments


def main():
    wb = openpyxl.load_workbook('/home/ubuntu/upload/AgendaProfissionais.xlsx', data_only=True)
    skip_sheets = {'Atendimentos', 'Cópia de Modelo'}

    all_appointments = []

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        professional_name = extract_professional_name(sheet_name)
        print(f"Processando: {professional_name}")

        appts = process_sheet(ws, professional_name)
        print(f"  -> {len(appts)} regras de recorrência encontradas")
        all_appointments.extend(appts)

    print(f"\nTotal de regras de recorrência: {len(all_appointments)}")

    # Verificar nomes suspeitos (horários no lugar de nomes)
    horario_pattern = re.compile(r'^\d{2}[h:]\d{2}')
    suspeitos = [a for a in all_appointments if horario_pattern.match(a['patient_full_name'].strip())]
    if suspeitos:
        print(f"\n[AVISO] {len(suspeitos)} nomes suspeitos encontrados:")
        for s in suspeitos:
            print(f"  {s['professional_full_name']} | {s['day_of_week_name']} {s['start_time']} | \"{s['patient_full_name']}\"")
    else:
        print("\n[OK] Nenhum nome suspeito encontrado.")

    # Salvar CSV
    output_path = '/home/ubuntu/agenda-aba/migration/recurring_appointments.csv'
    fieldnames = ['day_of_week', 'day_of_week_name', 'start_time', 'end_time',
                  'professional_full_name', 'patient_full_name',
                  'appointment_type', 'is_group_appointment', 'notes']

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_appointments)

    print(f"\nCSV salvo em: {output_path}")

    # Estatísticas
    professionals = set(a['professional_full_name'] for a in all_appointments)
    patients = set(a['patient_full_name'] for a in all_appointments)
    school = [a for a in all_appointments if a['appointment_type'] == 'school']
    dupla = [a for a in all_appointments if a['is_group_appointment'] == 'true']
    divergent = [a for a in all_appointments if 'divergente' in a['notes'].lower()]

    by_day = {}
    for a in all_appointments:
        d = a['day_of_week_name']
        by_day[d] = by_day.get(d, 0) + 1

    print(f"\n--- Estatísticas ---")
    print(f"Profissionais únicos: {len(professionals)}")
    print(f"Pacientes únicos: {len(patients)}")
    print(f"Atendimentos escolares: {len(school)}")
    print(f"Atendimentos em dupla: {len(dupla)}")
    print(f"Horários divergentes: {len(divergent)}")
    print(f"Por dia da semana:")
    for day in ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira']:
        print(f"  {day}: {by_day.get(day, 0)}")


if __name__ == '__main__':
    main()
