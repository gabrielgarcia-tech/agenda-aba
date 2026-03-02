"""
Script para processar a planilha AgendaProfissionais.xlsx e gerar o CSV de migração.
Lê todas as abas de profissionais, extrai os agendamentos e gera appointments.csv.
"""
import openpyxl
import csv
import re
from datetime import datetime, timedelta

# Mapeamento de dias da semana para número (0=segunda, 4=sexta)
DAYS = {
    'SEGUNDA-FEIRA': 0,
    'TERÇA-FEIRA': 1,
    'QUARTA-FEIRA': 2,
    'QUINTA-FEIRA': 3,
    'SEXTA-FEIRA': 4,
}

# Data de referência: próxima semana a partir de 03/03/2026 (segunda-feira)
BASE_DATE = datetime(2026, 3, 2)  # segunda-feira

# Mapeamento de horários para hora/minuto
def parse_horario(horario_str):
    """Converte '08h00 - 08h50' em (hora_inicio, minuto_inicio, hora_fim, minuto_fim)"""
    if not horario_str:
        return None
    match = re.match(r'(\d{2})h(\d{2})\s*-\s*(\d{2})h(\d{2})', str(horario_str))
    if match:
        return int(match.group(1)), int(match.group(2)), int(match.group(3)), int(match.group(4))
    return None

def get_datetime(day_offset, h_start, m_start, h_end, m_end):
    """Gera os datetimes de início e fim para um agendamento."""
    date = BASE_DATE + timedelta(days=day_offset)
    start = date.replace(hour=h_start, minute=m_start, second=0)
    end = date.replace(hour=h_end, minute=m_end, second=0)
    return start.strftime('%Y-%m-%d %H:%M:%S'), end.strftime('%Y-%m-%d %H:%M:%S')

def clean_name(name):
    """Remove espaços extras, sufixos de especialidade e normaliza o nome."""
    if not name:
        return None
    name = str(name).strip()
    # Remover sufixos como '- PSICO', '- FONO', '- TO', '- ABA'
    name = re.sub(r'\s*-\s*(PSICO|FONO|TO|ABA|PSICOLOGIA|FONOAUDIOLOGIA)\s*$', '', name, flags=re.IGNORECASE)
    return name.strip()

def extract_professional_name(sheet_name):
    """Extrai o nome do profissional do nome da aba."""
    # Remove a especialidade entre parênteses
    match = re.match(r'^(.+?)\s*[\(\(]', sheet_name)
    if match:
        return match.group(1).strip()
    return sheet_name.strip()

def process_cell(cell_value, horario_parsed, day_offset, professional_name):
    """
    Processa o conteúdo de uma célula e retorna uma lista de agendamentos.
    Retorna lista de dicts com os campos do CSV.
    """
    if not cell_value:
        return []
    
    value = str(cell_value).strip()
    
    # Ignorar células vazias, VAGO, SUPERVISÃO, Saída, etc.
    skip_keywords = ['VAGO', 'SUPERVISÃO', 'SUPERVISAO', 'Saída', 'Saida', 'FOLGA', 'FERIADO']
    if any(kw.lower() in value.lower() for kw in skip_keywords):
        return []
    
    if not horario_parsed:
        return []
    
    h_start, m_start, h_end, m_end = horario_parsed
    
    # Detectar horário divergente: "Das HHhMM às HHhMM Nome"
    divergent_match = re.match(r'Das?\s+(\d{1,2})h(\d{2})\s+[àa]s?\s+(\d{1,2})h(\d{2})\s*[-–]?\s*(.+)', value, re.IGNORECASE)
    if divergent_match:
        h_start = int(divergent_match.group(1))
        m_start = int(divergent_match.group(2))
        h_end = int(divergent_match.group(3))
        m_end = int(divergent_match.group(4))
        patient_name = clean_name(divergent_match.group(5))
        start_dt, end_dt = get_datetime(day_offset, h_start, m_start, h_end, m_end)
        return [{
            'start_time': start_dt,
            'end_time': end_dt,
            'professional_full_name': professional_name,
            'patient_full_name': patient_name,
            'appointment_type': 'clinic',
            'is_group_appointment': 'false',
            'notes': 'Horário divergente',
        }]
    
    # Detectar atendimento escolar: "Atendimento Escolar: Nome"
    school_match = re.match(r'Atendimento Escolar[:\s]+(.+)', value, re.IGNORECASE)
    if school_match:
        patient_name = clean_name(school_match.group(1))
        start_dt, end_dt = get_datetime(day_offset, h_start, m_start, h_end, m_end)
        return [{
            'start_time': start_dt,
            'end_time': end_dt,
            'professional_full_name': professional_name,
            'patient_full_name': patient_name,
            'appointment_type': 'school',
            'is_group_appointment': 'false',
            'notes': 'Atendimento Escolar',
        }]
    
    # Atendimento normal
    patient_name = clean_name(value)
    if not patient_name or len(patient_name) < 3:
        return []
    
    start_dt, end_dt = get_datetime(day_offset, h_start, m_start, h_end, m_end)
    return [{
        'start_time': start_dt,
        'end_time': end_dt,
        'professional_full_name': professional_name,
        'patient_full_name': patient_name,
        'appointment_type': 'clinic',
        'is_group_appointment': 'false',
        'notes': '',
    }]


def process_sheet(ws, professional_name):
    """Processa uma aba de profissional e retorna lista de agendamentos."""
    appointments = []
    
    # Encontrar a linha do cabeçalho (Horários, SEGUNDA-FEIRA, ...)
    header_row = None
    day_columns = {}  # col_index -> day_offset
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == 'SEGUNDA-FEIRA':
                header_row = cell.row
                # Mapear colunas para dias
                for c in ws.iter_cols(min_row=header_row, max_row=header_row, min_col=1, max_col=ws.max_column):
                    for hcell in c:
                        day_name = str(hcell.value).strip() if hcell.value else ''
                        if day_name in DAYS:
                            day_columns[hcell.column] = DAYS[day_name]
                break
        if header_row:
            break
    
    if not header_row or not day_columns:
        print(f"  [AVISO] Não encontrou cabeçalho em: {professional_name}")
        return []
    
    # Encontrar coluna de horários (geralmente coluna A ou B)
    horario_col = None
    for cell in ws[header_row]:
        if cell.value and 'horário' in str(cell.value).lower():
            horario_col = cell.column
            break
    if not horario_col:
        # Tentar coluna 1
        horario_col = 1
    
    # Processar linhas de dados
    for row in ws.iter_rows(min_row=header_row + 1):
        horario_cell = None
        for cell in row:
            if cell.column == horario_col:
                horario_cell = cell
                break
        
        if not horario_cell or not horario_cell.value:
            continue
        
        horario_str = str(horario_cell.value).strip()
        horario_parsed = parse_horario(horario_str)
        
        if not horario_parsed:
            continue  # Linha de resumo ou vazia
        
        # Processar cada coluna de dia
        for col_idx, day_offset in day_columns.items():
            cell_value = None
            for cell in row:
                if cell.column == col_idx:
                    cell_value = cell.value
                    break
            
            appts = process_cell(cell_value, horario_parsed, day_offset, professional_name)
            appointments.extend(appts)
    
    return appointments


def main():
    wb = openpyxl.load_workbook('/home/ubuntu/upload/AgendaProfissionais.xlsx', data_only=True)
    
    # Abas a ignorar
    skip_sheets = {'Atendimentos', 'Cópia de Modelo'}
    
    all_appointments = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        
        ws = wb[sheet_name]
        professional_name = extract_professional_name(sheet_name)
        print(f"Processando: {professional_name} ({sheet_name})")
        
        appts = process_sheet(ws, professional_name)
        print(f"  -> {len(appts)} agendamentos encontrados")
        all_appointments.extend(appts)
    
    print(f"\nTotal de agendamentos extraídos: {len(all_appointments)}")
    
    # Salvar CSV
    output_path = '/home/ubuntu/agenda-aba/migration/appointments.csv'
    fieldnames = ['start_time', 'end_time', 'professional_full_name', 'patient_full_name', 
                  'appointment_type', 'is_group_appointment', 'notes']
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_appointments)
    
    print(f"CSV salvo em: {output_path}")
    
    # Mostrar amostra
    print("\n--- Amostra dos primeiros 10 registros ---")
    for appt in all_appointments[:10]:
        print(appt)
    
    # Estatísticas
    professionals = set(a['professional_full_name'] for a in all_appointments)
    patients = set(a['patient_full_name'] for a in all_appointments)
    school = [a for a in all_appointments if a['appointment_type'] == 'school']
    
    print(f"\n--- Estatísticas ---")
    print(f"Profissionais únicos: {len(professionals)}")
    print(f"Pacientes únicos: {len(patients)}")
    print(f"Atendimentos escolares: {len(school)}")
    print(f"Atendimentos clínicos: {len(all_appointments) - len(school)}")


if __name__ == '__main__':
    main()
